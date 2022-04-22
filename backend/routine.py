import openpyxl


def append_to_routine(time, course_info, routine, day_name):
    if time == "08:30-10:00":
        routine[day_name][0].append(course_info)
    elif time == "10:00-11:30":
        routine[day_name][1].append(course_info)
    elif time == "11.30-01:00":
        routine[day_name][2].append(course_info)
    elif time == "01:00-2:30":
        routine[day_name][3].append(course_info)
    elif time == "02:30-4:00":
        routine[day_name][4].append(course_info)
    elif time == "04:00-05:30":
        routine[day_name][5].append(course_info)


def get_simplified_routine(course_codes=[], master_routine="routine.xlsx", sheet=None):
    """search through the master routine and generate a simplified routine

    Returns:
        python dict: simplified routine
    """

    workbook = openpyxl.load_workbook(master_routine)

    sheet = workbook[sheet]

    routine = {
        "Saturday": [[], [], [], [], [], []],
        "Sunday": [[], [], [], [], [], []],
        "Monday": [[], [], [], [], [], []],
        "Tuesday": [[], [], [], [], [], []],
        "Wednesday": [[], [], [], [], [], []],
        "Thursday": [[], [], [], [], [], []],
    }

    day = None
    start_row = 6
    end_row = 104

    cols_with_course_codes = [2, 5, 8, 11, 14, 17]
    rows_with_day_names = [87, 70, 54, 38, 22, 6]  # column is always 1
    times = ["08:30-10:00", "10:00-11:30", "11.30-01:00",
             "01:00-2:30", "02:30-4:00", "04:00-05:30"]

    # traversing columns with sub names from left-right
    for time, a_col_with_course_code in zip(times, cols_with_course_codes):

        # setting counter = start_row every time we finish a column & move to next time frame
        counter = start_row

        while counter < end_row:  # scan one time frame to see if there is a class in that time

            # the course code at the current point of the traversing
            course_code = sheet.cell(
                row=counter, column=a_col_with_course_code).value
            # the room no. at the current point of the traversing
            room_no = sheet.cell(
                row=counter, column=a_col_with_course_code - 1).value
            # the teacher initials at the current point of the traversing
            teacher_initials = sheet.cell(
                row=counter, column=a_col_with_course_code + 1).value

            # this whole block is to find the day, so we go up until we find the day
            i = counter  # i is just an increment variable
            while i + 1 not in rows_with_day_names:
                if i in rows_with_day_names:
                    day = i
                    break
                i -= 1
            # find day block end

            # here (6, 22, 38, 54, 70, 87) each represents a row that has day name. so for ex, 6 corresponds to Saturday and so on. this is according to the master routine xlsx, they are always constant
            if course_code in course_codes:  # if we have the current course this semester
                course_info = {"time": time, "course_code": course_code,
                               "room_no": room_no, "teacher_initials": teacher_initials}
                if day == 6:
                    day_name = "Saturday"
                    append_to_routine(time, course_info, routine, day_name)
                elif day == 22:
                    day_name = "Sunday"
                    append_to_routine(time, course_info, routine, day_name)
                elif day == 38:
                    day_name = "Monday"
                    append_to_routine(time, course_info, routine, day_name)
                elif day == 54:
                    day_name = "Tuesday"
                    append_to_routine(time, course_info, routine, day_name)
                elif day == 70:
                    day_name = "Wednesday"
                    append_to_routine(time, course_info, routine, day_name)
                elif day == 87:
                    day_name = "Thursday"
                    append_to_routine(time, course_info, routine, day_name)

            counter += 1

    return routine
