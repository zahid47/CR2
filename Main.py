import datetime
import openpyxl
from pprint import pprint

# opening an excel file and selecting the first sheet which is named "Sheet1"
workbook = openpyxl.load_workbook("sample_routine.xlsx")
sheet = workbook["Sheet1"]

# dictionary that contains sub codes, key = semester no., value = list of sub codes
subjects = {
    4: ["SE211", "SE212", "SE213", "SE214", "SE215", "CS211"],
    5: ["SE221", "SE223", "SE224", "SE225", "SE226", "SE222"]
}


def get_day(day_number):
    if day_number == 0:
        return "Monday"
    elif day_number == 1:
        return "Tuesday"
    elif day_number == 2:
        return "Wednesday"
    elif day_number == 3:
        return "Thursday"
    elif day_number == 4:
        return "Friday"
    elif day_number == 5:
        return "Saturday"
    elif day_number == 6:
        return "Sunday"


def get_day_number(day):
    if day == "Monday":
        return 0
    elif day == "Tuesday":
        return 1
    elif day == "Wednesday":
        return 2
    elif day == "Thursday":
        return 3
    elif day == "Friday":
        return 4
    elif day == "Saturday":
        return 5
    elif day == "Sunday":
        return 6


def get_subjects_with_section(subjects_list, semester, section):
    section = section.upper()
    subjects_with_section = []
    for subject in subjects_list[semester]:
        subjects_with_section.append(subject + section)
    return subjects_with_section


def get_routine(semester, section, want_tomorrow=False, cs_major=False, day=get_day(datetime.datetime.now().weekday())):
    result = []
    section = section.upper()
    routine = {}
    # no_of_subs = []

    if want_tomorrow:  # == True
        day = get_day(datetime.datetime.now().weekday() + 1)

    tomorrow = get_day(get_day_number(day) + 1)

    # print(f"Today = {day}")  # debug
    # print(f"Tomorrow = {tomorrow}")  # debug

    row_length = len(sheet["A"])
    col_length = len(sheet[1])

    # print(row_length)  # debug
    # print(col_length)  # debug

    for row in range(1, row_length):  # traversing rows from top-down

        if sheet.cell(row=row, column=1).value == day:  # we found the day! also, day name always in first column(col=1)

            while sheet.cell(row=row, column=1).value != tomorrow:  # stop when we reach next day!

                # now we are inside today's routine in the excel file

                for column in range(2, col_length, 3):
                    sub = sheet.cell(row=row, column=column).value

                    if sub in get_subjects_with_section(subjects, semester, section):
                        # no_of_subs.append(sub)

                        time = sheet.cell(row=4, column=column - 1).value
                        room = sheet.cell(row=row, column=column - 1).value
                        teacher = sheet.cell(row=row, column=column + 1).value

                        result.append(f"Time: {time}, Subject: {sub}, Room: {room}, Teacher: {teacher}")

                        '''
                        for sub_serial in range(len(no_of_subs)):
                            routine[sub_serial] = {}
                            routine[sub_serial]["Subject_code"] = sub
                            routine[sub_serial]["Room"] = sheet.cell(row=row, column=column - 1).value
                            routine[sub_serial]["Teacher"] = sheet.cell(row=row, column=column + 1).value
                            routine[sub_serial]["Time"] = sheet.cell(row=4, column=column - 1).value '''

                row += 1

    '''
    if len(routine) == 0:
        return f"No classes in {day}!"
    return routine'''

    if len(result) == 0:
        return f"No classes in {day}!"
    return result


no_of_classes = len(get_routine(4, "c"))

print(f"You have {no_of_classes} class(es) in {get_day(datetime.datetime.now().weekday())}.\n")

for entry in get_routine(4, "d"):
    print(entry)
